#!/usr/bin/env python3
"""Gera Excel de resumo financeiro a partir da estrutura da tela (hierarquia)."""

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def money(v: float) -> str:
    neg = v < 0
    s = f"R$ {abs(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"-{s}" if neg else s


class RowCursor:
    __slots__ = ("n",)

    def __init__(self, start: int) -> None:
        self.n = start


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo financeiro"

    header_fill = PatternFill("solid", fgColor="E8E8E8")
    section_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    red = Font(color="C00000")

    COL_LEFT_LABEL = 1
    COL_RIGHT_LABEL = 4
    LAST_COL_BAR = 5

    ws["A1"] = "Resumo financeiro"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    ws["A2"] = "Período (exemplo da tela)"
    ws["B2"] = "Hoje"

    # Barra "Total atual" a largura completa (linha 4)
    bar_row = 4
    ws.cell(bar_row, COL_LEFT_LABEL, "Total atual")
    ws.cell(bar_row, COL_LEFT_LABEL + 1, money(2200.0))
    for c in range(1, LAST_COL_BAR + 1):
        cell = ws.cell(bar_row, c)
        cell.fill = header_fill
        cell.font = bold
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for c in range(COL_LEFT_LABEL + 2, LAST_COL_BAR + 1):
        ws.cell(bar_row, c, None)

    start_blocks = bar_row + 1
    cur_left = RowCursor(start_blocks)
    cur_right = RowCursor(start_blocks)

    def write_line(
        cursor: RowCursor,
        col_label: int,
        indent: int,
        label: str,
        value: Optional[str] = None,
        *,
        is_header_bar: bool = False,
        is_section: bool = False,
        negative: bool = False,
    ) -> None:
        r = cursor.n
        ws.cell(r, col_label, "  " * indent + label)
        col_val = col_label + 1
        if value is not None:
            ws.cell(r, col_val, value)
        cell_a = ws.cell(r, col_label)
        cell_b = ws.cell(r, col_val) if value is not None else None
        for c in (cell_a, cell_b):
            if c is None:
                continue
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
        if is_header_bar or is_section:
            row_fill = header_fill if is_header_bar else section_fill
            cell_a.fill = row_fill
            cell_a.font = bold
            if cell_b:
                cell_b.fill = row_fill
                cell_b.font = (
                    Font(bold=True, color="C00000") if negative else bold
                )
        elif negative and cell_b:
            cell_b.font = red
        cursor.n += 1

    # --- Coluna esquerda: Disponível p/ saque (saldo disponível) ---
    write_line(
        cur_left,
        COL_LEFT_LABEL,
        0,
        "Disponível p/ saque",
        money(1800.0),
        is_section=True,
    )
    write_line(cur_left, COL_LEFT_LABEL, 1, "Saldo final", money(1800.0))

    write_line(
        cur_left, COL_LEFT_LABEL, 1, "Entrada", money(2000.0), is_section=True
    )
    write_line(
        cur_left,
        COL_LEFT_LABEL,
        2,
        "Venda (líquido)",
        money(1200.0),
        is_section=True,
    )
    write_line(cur_left, COL_LEFT_LABEL, 3, "Valor Bruto", money(1250.0))
    write_line(cur_left, COL_LEFT_LABEL, 3, "MDR", money(-10.0), negative=True)
    write_line(cur_left, COL_LEFT_LABEL, 3, "Juros", money(-40.0), negative=True)

    write_line(
        cur_left,
        COL_LEFT_LABEL,
        2,
        "Link (líquido)",
        money(500.0),
        is_section=True,
    )
    write_line(cur_left, COL_LEFT_LABEL, 3, "Valor Bruto", money(525.0))
    write_line(cur_left, COL_LEFT_LABEL, 3, "MDR", money(-10.0), negative=True)
    write_line(cur_left, COL_LEFT_LABEL, 3, "Juros", money(-15.0), negative=True)

    write_line(cur_left, COL_LEFT_LABEL, 2, "Reembolso", money(250.0))
    write_line(cur_left, COL_LEFT_LABEL, 2, "Depósito", money(350.0))

    write_line(
        cur_left,
        COL_LEFT_LABEL,
        1,
        "Saídas",
        money(-500.0),
        is_section=True,
        negative=True,
    )
    write_line(
        cur_left,
        COL_LEFT_LABEL,
        2,
        "Contestação (líquido)",
        money(-50.0),
        is_section=True,
        negative=True,
    )
    write_line(
        cur_left,
        COL_LEFT_LABEL,
        3,
        "Valor da contestação",
        money(-55.0),
        negative=True,
    )
    write_line(cur_left, COL_LEFT_LABEL, 3, "Devolução de MDR", money(5.0))

    write_line(
        cur_left,
        COL_LEFT_LABEL,
        2,
        "Estorno",
        money(-100.0),
        is_section=True,
        negative=True,
    )
    write_line(
        cur_left, COL_LEFT_LABEL, 3, "Valor Bruto", money(-110.0), negative=True
    )
    write_line(cur_left, COL_LEFT_LABEL, 3, "Devolução de MDR", money(5.0))
    write_line(cur_left, COL_LEFT_LABEL, 3, "Devolução de juros", money(5.0))

    write_line(
        cur_left,
        COL_LEFT_LABEL,
        2,
        "Transferência (bruto)",
        money(-350.0),
        is_section=True,
        negative=True,
    )
    write_line(
        cur_left,
        COL_LEFT_LABEL,
        3,
        "Valor transferido",
        money(-349.01),
        negative=True,
    )
    write_line(
        cur_left,
        COL_LEFT_LABEL,
        3,
        "Taxa de transferência",
        money(-0.99),
        negative=True,
    )

    write_line(cur_left, COL_LEFT_LABEL, 1, "Saldo inicial", money(200.0))

    # --- Coluna direita: Lançamentos futuros (alinhado ao topo com o bloco esquerdo) ---
    write_line(
        cur_right,
        COL_RIGHT_LABEL,
        0,
        "Lançamentos futuros",
        money(300.0),
        is_section=True,
    )

    write_line(
        cur_right, COL_RIGHT_LABEL, 1, "Entrada", money(500.0), is_section=True
    )
    write_line(
        cur_right,
        COL_RIGHT_LABEL,
        2,
        "Venda (líquido)",
        money(400.0),
        is_section=True,
    )
    write_line(cur_right, COL_RIGHT_LABEL, 3, "Valor Bruto", money(420.0))
    write_line(cur_right, COL_RIGHT_LABEL, 3, "MDR", money(-10.0), negative=True)
    write_line(cur_right, COL_RIGHT_LABEL, 3, "Juros", money(-10.0), negative=True)

    write_line(
        cur_right,
        COL_RIGHT_LABEL,
        2,
        "Link (líquido)",
        money(100.0),
        is_section=True,
    )
    write_line(cur_right, COL_RIGHT_LABEL, 3, "Valor Bruto", money(105.0))
    write_line(cur_right, COL_RIGHT_LABEL, 3, "MDR", money(-2.0), negative=True)
    write_line(cur_right, COL_RIGHT_LABEL, 3, "Juros", money(-3.0), negative=True)

    write_line(
        cur_right,
        COL_RIGHT_LABEL,
        1,
        "Saídas",
        money(-200.0),
        is_section=True,
        negative=True,
    )
    write_line(
        cur_right,
        COL_RIGHT_LABEL,
        2,
        "Contestação (líquido)",
        money(-50.0),
        is_section=True,
        negative=True,
    )
    write_line(
        cur_right,
        COL_RIGHT_LABEL,
        3,
        "Valor da contestação",
        money(-55.0),
        negative=True,
    )
    write_line(cur_right, COL_RIGHT_LABEL, 3, "Devolução de MDR", money(5.0))

    write_line(
        cur_right,
        COL_RIGHT_LABEL,
        2,
        "Estorno",
        money(-150.0),
        is_section=True,
        negative=True,
    )
    write_line(
        cur_right,
        COL_RIGHT_LABEL,
        3,
        "Valor Bruto",
        money(-170.0),
        negative=True,
    )
    write_line(cur_right, COL_RIGHT_LABEL, 3, "Devolução de MDR", money(5.0))
    write_line(cur_right, COL_RIGHT_LABEL, 3, "Devolução de juros", money(15.0))

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 16

    out_dir = Path(__file__).resolve().parent / "Gestão financeira"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / "Resumo_financeiro.xlsx"
    wb.save(out)
    print(f"Salvo: {out}")


if __name__ == "__main__":
    main()
